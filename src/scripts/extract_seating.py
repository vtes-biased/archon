import openpyxl
import json

wb = openpyxl.load_workbook("thearchon1.5l.xlsx", data_only=True)
ws = wb["Optimal Seating 3R+F"]


def chunk_into_tables(players: list[int], num_players: int) -> list[list[int]]:
    """Split flat list into 5-player and 4-player tables."""
    # Formula: num_4_tables = (5 - (N % 5)) % 5
    num_4_tables = (5 - (num_players % 5)) % 5
    num_5_tables = (num_players - 4 * num_4_tables) // 5
    tables = []
    idx = 0
    for _ in range(num_5_tables):
        tables.append(players[idx : idx + 5])
        idx += 5
    for _ in range(num_4_tables):
        tables.append(players[idx : idx + 4])
        idx += 4
    return tables


def parse_round_row(ws, row, num_players: int):
    # First collect all values
    all_values = []
    col = 6
    while col <= 500:
        val = ws.cell(row=row, column=col).value
        if val is None:
            col += 1
            continue
        if isinstance(val, (int, float)):
            all_values.append(int(val))
        col += 1
        if len(all_values) >= num_players:
            break
    # Chunk into proper tables
    return chunk_into_tables(all_values, num_players)


data = {}
row = 7

while row <= ws.max_row:
    cell_d = ws.cell(row=row, column=4).value
    cell_e = ws.cell(row=row, column=5).value

    if isinstance(cell_d, int) and cell_e and "Players" in str(cell_e):
        num_players = cell_d
        rounds = []

        for r in range(1, 4):
            round_row = row + r
            round_label = ws.cell(row=round_row, column=5).value
            if round_label and "Round" in str(round_label):
                tables = parse_round_row(ws, round_row, num_players)
                rounds.append(tables)

        if len(rounds) == 3:
            data[num_players] = rounds
        row += 5
    else:
        row += 1

# Verify
for n in [4, 5, 16, 20]:
    if n in data:
        print(f"{n} players:")
        for i, r in enumerate(data[n]):
            print(f"  Round {i + 1}: {r}")

print(f"\nTotal: {len(data)} player counts")

with open("src/front/optimal_seating_3r.json", "w") as f:
    json.dump(data, f, separators=(",", ":"))
print("Saved to src/front/optimal_seating_3r.json")

